import openpyxl
from openpyxl.styles import PatternFill


def getRowCount(filepath, sheetname):
    workbook = openpyxl.load_workbook(filepath)
    sheet = workbook[sheetname]
    row = sheet.max_row
    return row


def getColumnCount(filepath, sheetname):
    workbook = openpyxl.load_workbook(filepath)
    sheet = workbook[sheetname]
    column = sheet.max_column
    return column


def readData(filepath, sheetname, rowNo, columnNo):
    workbook = openpyxl.load_workbook(filepath)
    sheet = workbook[sheetname]
    data = sheet.cell(rowNo, columnNo).value
    return data


def writeData(filepath, sheetname, rowNo, columnNo, data):
    workbook = openpyxl.load_workbook(filepath)
    sheet = workbook[sheetname]
    sheet.cell(rowNo, columnNo).value = data
    workbook.save(filepath)


def fillGreenColor(filepath, sheetname, rowNo, columnNo):
    workbook = openpyxl.load_workbook(filepath)
    sheet = workbook[sheetname]
    greenFill = PatternFill(start_color='60b212',
                            end_color='60b212',
                            fill_type='solid', )
    sheet.cell(rowNo, columnNo).fill = greenFill
    workbook.save(filepath)


def fillRedColor(filepath, sheetname, rowNo, columnNo):
    workbook = openpyxl.load_workbook(filepath)
    sheet = workbook[sheetname]
    redFill = PatternFill(start_color='ff0000',
                          end_color='ff0000',
                          fill_type='solid', )
    sheet.cell(rowNo, columnNo).fill = redFill
    workbook.save(filepath)
